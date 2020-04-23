# -*- coding:utf-8 -*-
import openpyxl as op
from platform import system as psys
from os import system as osys,getcwd,path,remove
import configparser as cp
import datetime as dt

ps='\\' if psys()=='Windows' else '/'

class struct:
    def __init__(self):
        c=cp.ConfigParser()
        c.read(path.dirname(path.realpath(__file__))+'/config.ini',encoding='utf-8')
        self.subject=c['DEFAULT']['subject']
        self.outputFileName=c['DEFAULT']['outputFileName']
        self.templateName=c['DEFAULT']['templateName']
        self.templateSheet=c['DEFAULT']['templateSheet']
        self.correct=c['DEFAULT']['correct']
        self.title=c['DEFAULT']['title']

err=list()

def is_Chinese(ch):
    if '\u4e00' <= ch <= '\u9fff':
        return True
    return False

def func():
    c=struct()
    wb=op.load_workbook(c.templateName)
    ws=wb[c.templateSheet]
    wb.close()
    sid=dict()
    i='A'
    ws['A1']=c.title.format(month=dt.datetime.now().strftime("%m"),day=dt.datetime.now().strftime("%d"))
    for cell in ws[2]:
        sid[cell.value]=i
        i=chr(ord(i)+1)
    nid,i=dict(),0
    for name in ws[sid['姓名']]:
        i+=1
        if i<=2:
            continue
        nid[name.value]=i
    fp=getcwd()+ps+'input.txt'
    inf=open(fp,'w+')
    inf.close()
    print('''请打开%s，并将内容输入到该文件中。
一般来说，第一行应该有一个@。
如果数据有序号，请在@后面加一个I；
如果数据有附加信息，请在@后面加一个C。
（两者可同时出现，不区分大小写。如果结果为@I/@i，可省略不输入）
如果以上皆无，第一行应该有且只有一个@。
输入完成后，'''%fp)
    nc,ni=True,True
    if psys()=="Windows":
        osys("pause")
    else:
        print('请在此输入任何内容并换行以继续...')
        input()
    with open(fp,'r',encoding='utf-8') as f:
        mstr=f.readline().strip()
        if mstr and mstr[0]=='@':
            mstr=mstr.upper()
            if 'C' in mstr:
                nc=False
            if 'I' in mstr:
                ni=False
            mstr=f.readline().strip()
        else:
            ni=False
        while mstr:
            try:
                if mstr=='':
                    continue
                if nc and ni:
                    ws['%s%d'%(sid[c.subject],nid[mstr])]=c.correct
                else:
                    ns=0
                    if not ni:
                        ns=-1
                        for i in range(0,len(mstr)):
                            ch=mstr[i]
                            if is_Chinese(ch):
                                if ns==-1:
                                    ns=i
                                    break
                    else:
                        ns=0
                    if not nc:
                        ne=-1
                        for i in range(1,len(mstr)+1):
                            ch=mstr[-i]
                            if is_Chinese(ch):
                                if ne==-1:
                                    ne=-i+1
                                    break
                    else:
                        ne=len(mstr)
                    ws['%s%d'%(sid[c.subject],nid[mstr[ns:ne]])]=c.correct if nc else mstr[ne:]
                    pass
            except:
                err.append(mstr)
            mstr=f.readline().strip()
    if path.exists(fp):
        remove(fp)
    wb.save(c.outputFileName)
    print('文件已输出到%s'%(getcwd()+ps+c.outputFileName))
    if len(err)!=0:
        print('以下数据未能成功处理：')
        for mstr in err:
            print(mstr)

if __name__=='__main__':
    try:
        func()
    except KeyboardInterrupt:
        pass
    except Exception as e:
        print('程序出错')
        print(e)

    if psys()=="Windows":
        osys('pause')
